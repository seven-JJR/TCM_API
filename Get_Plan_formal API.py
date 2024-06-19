import openpyxl
import datetime
import random

excel=openpyxl.load_workbook(r'C:\excel_to_TCMAPI.xlsx')
basicinfo_sheet = excel['test_plan']
getplan_sheet=excel['test_case']
schedule_start=basicinfo_sheet['B4'].value
schedule_end=basicinfo_sheet['B5'].value
schedule_start=schedule_start.split(",")#把开始日期变成列表，方便后面使用
schedule_end=schedule_end.split(",")#把结束日期变成列表，方便后面使用
getplan_sheet_max_row=getplan_sheet.max_row# 获得最大行数

start_date_format = datetime.date(int(schedule_start[0]), int(schedule_start[1]), int(schedule_start[2])) #开始日期变为时间格式，方便后面比较时间前后大小
end_date_format = datetime.date(int(schedule_end[0]), int(schedule_end[1]), int(schedule_end[2]))#开始日期变为时间格式，方便后面比较时间前后大小

start_time_hour_am_list=['8','9','10','11','13','14','15','16','18','19']
date_format = '%Y-%m-%d %H:%M'
date_format_1 = '%H:%M'
for every_row in range(2,getplan_sheet_max_row+1):
    start_time_hour_am = str(random.randint(8, 11))
    start_time_hour_minutes_1 = str(random.randint(0, 59))
    start_time_hour_minutes_2 = str(random.randint(31, 59))
    every_row_list = []  # 给每一行建立一个空列表来装每行一具体的值
    for cell in getplan_sheet[every_row]:  # caseinfo_sheet[i] i是就代表caseinfo sheet的第几行，cell是代表这一行中的每一个单元格
         every_row_list.append(cell.value)
    if start_date_format.weekday() == 5:  #如果时间等于周六， .weekday()返回整数（0到6）对于周1到周天
        excel_start_date=start_date_format + datetime.timedelta(days=2)# 如果等于周六时间就后推2天，从周一开始，跳过周末
        if excel_start_date <= end_date_format: #加2天后如果时间小于截止日就继续下面的步骤
            getplan_sheet.cell(column=5, row=every_row, value=excel_start_date)#填开始测试日期，start date
            start_date_format=excel_start_date+datetime.timedelta(days=1)#填完后时间后推一天
            if every_row_list[9] <= 480: #如果laoding 小于等于480分钟的，就从当天的8到11点开始测试，目的是尽量避免到凌晨测完
                if start_time_hour_am !='8': # 如果不是8点开始，分钟就可以是0到59
                    excel_start_time=start_time_hour_am+":"+start_time_hour_minutes_1#开始时间的字符串格式
                    excel_start_time_format=datetime.datetime.strptime(excel_start_time, date_format_1)#再转为时间格式
                    excel_start_time_str=excel_start_time_format.strftime(date_format_1)#再转为符合TCM格式的字符串
                    getplan_sheet.cell(column=7, row=every_row, value=excel_start_time_str)#填开始测试时间，start time
                    start_datetime=excel_start_date.strftime("%Y-%m-%d")+" "+excel_start_time #先得出完整开始测试时间的字符串格式，例2023-11-1 1:2 一点2分，这是为了根据workloading推算截止日
                    start_datetime_timeformat=datetime.datetime.strptime(start_datetime,date_format)#开始时间字符串格式转时间格式
                    end_datetime=start_datetime_timeformat+datetime.timedelta(minutes=every_row_list[9])#获取完整年月日几点几分的结束日期等于加上workloading后的,结束时间是时间格式的，为了填excel，还要转字符串格式
                    end_date_excel=end_datetime.date()#end_datetime.date()返回时间格式的年月日，可以直接填到excel,无需转为字符串
                    getplan_sheet.cell(column=6, row=every_row, value=end_date_excel)  # 填结束测试日期，end date
                    end_time_excel=(str(end_datetime.time()))[:-3]#虽然date_format = '%Y-%m-%d %H:%M'，end_datetime.time()还是返回的时间带有秒数，例2023-11-1 1:2:23 且为时间格式
                    # 所以再转字符串并且去除秒的部分1:2:23不要后面3位
                    getplan_sheet.cell(column=8, row=every_row, value=end_time_excel)#填结束测试时间，end time,到此为止，时间全部填完，开始时间和结束时间填的字符串格式，开始日期和结束日期填的时间格式
                    # if every_row_list[10] is not  None:#如果有issue end time就延长.
                    #     DFT_list=every_row_list[10].split('\n')
                    #     length_DFT_list=len(DFT_list)
                    #     getplan_sheet.cell(column=7, row=every_row, value='Issue verify')
                    #     getplan_sheet.cell(column=8, row=every_row, value=str(0.5*length_DFT_list))


                else: #8点开始，分钟就只能是31到59
                    excel_start_time = start_time_hour_am + ":" + start_time_hour_minutes_2#8点开始，分钟就只能是31到59，#开始时间的字符串格式
                    excel_start_time_format = datetime.datetime.strptime(excel_start_time, date_format_1)  # 再转为时间格式
                    excel_start_time_str = excel_start_time_format.strftime(date_format_1)  # 再转为符合TCM格式的字符串
                    getplan_sheet.cell(column=7, row=every_row, value=excel_start_time_str)  # 填开始测试时间，start time
                    start_datetime = excel_start_date.strftime("%Y-%m-%d") + " " + excel_start_time  # 先得出完整开始测试时间的字符串格式，例2023-11-1 1:2 一点2分，这是为了根据workloading推算截止日
                    start_datetime_timeformat = datetime.datetime.strptime(start_datetime,date_format)  # 开始时间字符串格式转时间格式
                    end_datetime = start_datetime_timeformat + datetime.timedelta(minutes=every_row_list[9])  # 获取完整年月日几点几分的结束日期等于加上workloading后的,结束时间是时间格式的，为了填excel，还要转字符串格式
                    end_date_excel = end_datetime.date()  # end_datetime.date()返回时间格式的年月日，可以直接填到excel,无需转为字符串
                    getplan_sheet.cell(column=6, row=every_row, value=end_date_excel)  # 填结束测试日期，end date
                    end_time_excel = (str(end_datetime.time()))[:-3]  # 虽然date_format = '%Y-%m-%d %H:%M'，end_datetime.time()还是返回的时间带有秒数，例2023-11-1 1:2:23 且为时间格式
                    # 所以再转字符串并且去除秒的部分1:2:23不要后面3位
                    getplan_sheet.cell(column=8, row=every_row,value=end_time_excel)  # 填结束测试时间，end time,到此为止，时间全部填完，开始时间和结束时间填的字符串格式，开始日期和结束日期填的时间格式
                    # if every_row_list[10] is not  None:
                    #     DFT_list=every_row_list[10].split('\n')
                    #     length_DFT_list=len(DFT_list)
                    #     getplan_sheet.cell(column=7, row=every_row, value='Issue verify')
                    #     getplan_sheet.cell(column=8, row=every_row, value=str(0.5*length_DFT_list))


                    # 而且恰好开始结束时间是TCM标准的字符串格式, 开始结束日期不是
            else:# 大于480分钟的随机生成开始测试时间
                start_time_hour_am = start_time_hour_am_list[random.randint(0,9)]#列表中的合理的时间中，任意时随机间开始
                if start_time_hour_am != '8': # 如果不是8点开始，分钟就可以是0到59
                    excel_start_time = start_time_hour_am + ":" + start_time_hour_minutes_1  # 开始时间的字符串格式
                    excel_start_time_format = datetime.datetime.strptime(excel_start_time, date_format_1)  # 再转为时间格式
                    excel_start_time_str = excel_start_time_format.strftime(date_format_1)  # 再转为符合TCM格式的字符串
                    getplan_sheet.cell(column=7, row=every_row, value=excel_start_time_str)  # 填开始测试时间，start time
                    start_datetime = excel_start_date.strftime("%Y-%m-%d") + " " + excel_start_time  # 先得出完整开始测试时间的字符串格式，例2023-11-1 1:2 一点2分，这是为了根据workloading推算截止日
                    start_datetime_timeformat = datetime.datetime.strptime(start_datetime,date_format)  # 开始时间字符串格式转时间格式
                    end_datetime = start_datetime_timeformat + datetime.timedelta(minutes=every_row_list[9])  # 获取完整年月日几点几分的结束日期等于加上workloading后的,结束时间是时间格式的，为了填excel，还要转字符串格式
                    end_date_excel = end_datetime.date()  # end_datetime.date()返回时间格式的年月日，可以直接填到excel,无需转为字符串
                    getplan_sheet.cell(column=6, row=every_row, value=end_date_excel)  # 填结束测试日期，end date
                    end_time_excel = (str(end_datetime.time()))[:-3]  # 虽然date_format = '%Y-%m-%d %H:%M'，end_datetime.time()还是返回的时间带有秒数，例2023-11-1 1:2:23 且为时间格式
                    # 所以再转字符串并且去除秒的部分1:2:23不要后面3位
                    getplan_sheet.cell(column=8, row=every_row,value=end_time_excel)  # 填结束测试时间，end time,到此为止，时间全部填完，开始时间和结束时间填的字符串格式，开始日期和结束日期填的时间格式
                    # if every_row_list[10] is not  None:
                    #     DFT_list=every_row_list[10].split('\n')
                    #     length_DFT_list=len(DFT_list)
                    #     getplan_sheet.cell(column=7, row=every_row, value='Issue verify')
                    #     getplan_sheet.cell(column=8, row=every_row, value=str(0.5*length_DFT_list))

                    # 而且恰好结束时间是TCM标准的字符串格式
                else:#8点开始，分钟就只能是31到59
                    excel_start_time = start_time_hour_am + ":" + start_time_hour_minutes_2  # 8点开始，分钟就只能是31到59，#开始时间的字符串格式
                    excel_start_time_format = datetime.datetime.strptime(excel_start_time, date_format_1)  # 再转为时间格式
                    excel_start_time_str = excel_start_time_format.strftime(date_format_1)  # 再转为符合TCM格式的字符串
                    getplan_sheet.cell(column=7, row=every_row, value=excel_start_time_str)  # 填开始测试时间，start time
                    start_datetime = excel_start_date.strftime("%Y-%m-%d") + " " + excel_start_time  # 先得出完整开始测试时间的字符串格式，例2023-11-1 1:2 一点2分，这是为了根据workloading推算截止日
                    start_datetime_timeformat = datetime.datetime.strptime(start_datetime,date_format)  # 开始时间字符串格式转时间格式
                    end_datetime = start_datetime_timeformat + datetime.timedelta(minutes=every_row_list[9])  # 获取完整年月日几点几分的结束日期等于加上workloading后的,结束时间是时间格式的，为了填excel，还要转字符串格式
                    end_date_excel = end_datetime.date()  # end_datetime.date()返回时间格式的年月日，可以直接填到excel,无需转为字符串
                    getplan_sheet.cell(column=6, row=every_row, value=end_date_excel)  # 填结束测试日期，end date
                    end_time_excel = (str(end_datetime.time()))[:-3]  # 虽然date_format = '%Y-%m-%d %H:%M'，end_datetime.time()还是返回的时间带有秒数，例2023-11-1 1:2:23 且为时间格式
                    # 所以再转字符串并且去除秒的部分1:2:23不要后面3位
                    getplan_sheet.cell(column=8, row=every_row,value=end_time_excel)
                    # if every_row_list[10] is not  None:
                    #     DFT_list=every_row_list[10].split('\n')
                    #     length_DFT_list=len(DFT_list)
                    #     getplan_sheet.cell(column=7, row=every_row, value='Issue verify')
                    #     getplan_sheet.cell(column=8, row=every_row, value=str(0.5*length_DFT_list))

        else:#加2天后如果时间大于截止日，case开始测试日期就重置到schedule 开始时间再开始循环
            excel_start_date=datetime.date(int(schedule_start[0]), int(schedule_start[1]), int(schedule_start[2]))#加2天后如果时间大于截止日，case开始测试日期就重置到schedule 开始时间再开始循环
            getplan_sheet.cell(column=5, row=every_row, value=excel_start_date)
            start_date_format=excel_start_date+datetime.timedelta(days=1)
            if every_row_list[9] <= 480:
                if start_time_hour_am != '8':
                    excel_start_time = start_time_hour_am + ":" + start_time_hour_minutes_1  # 开始时间的字符串格式
                    excel_start_time_format = datetime.datetime.strptime(excel_start_time, date_format_1)  # 再转为时间格式
                    excel_start_time_str = excel_start_time_format.strftime(date_format_1)  # 再转为符合TCM格式的字符串
                    getplan_sheet.cell(column=7, row=every_row, value=excel_start_time_str)  # 填开始测试时间，start time
                    start_datetime = excel_start_date.strftime("%Y-%m-%d") + " " + excel_start_time  # 先得出完整开始测试时间的字符串格式，例2023-11-1 1:2 一点2分，这是为了根据workloading推算截止日
                    start_datetime_timeformat = datetime.datetime.strptime(start_datetime,date_format)  # 开始时间字符串格式转时间格式
                    end_datetime = start_datetime_timeformat + datetime.timedelta(minutes=every_row_list[9])  # 获取完整年月日几点几分的结束日期等于加上workloading后的,结束时间是时间格式的，为了填excel，还要转字符串格式
                    end_date_excel = end_datetime.date()  # end_datetime.date()返回时间格式的年月日，可以直接填到excel,无需转为字符串
                    getplan_sheet.cell(column=6, row=every_row, value=end_date_excel)  # 填结束测试日期，end date
                    end_time_excel = (str(end_datetime.time()))[ :-3]  # 虽然date_format = '%Y-%m-%d %H:%M'，end_datetime.time()还是返回的时间带有秒数，例2023-11-1 1:2:23 且为时间格式
                    # 所以再转字符串并且去除秒的部分1:2:23不要后面3位
                    getplan_sheet.cell(column=8, row=every_row, value=end_time_excel)
                    # if every_row_list[10] is not  None:
                    #     DFT_list=every_row_list[10].split('\n')
                    #     length_DFT_list=len(DFT_list)
                    #     getplan_sheet.cell(column=7, row=every_row, value='Issue verify')
                    #     getplan_sheet.cell(column=8, row=every_row, value=str(0.5*length_DFT_list))

                else:
                    excel_start_time = start_time_hour_am + ":" + start_time_hour_minutes_2  # 8点开始，分钟就只能是31到59，#开始时间的字符串格式
                    excel_start_time_format = datetime.datetime.strptime(excel_start_time, date_format_1)  # 再转为时间格式
                    excel_start_time_str = excel_start_time_format.strftime(date_format_1)  # 再转为符合TCM格式的字符串
                    getplan_sheet.cell(column=7, row=every_row, value=excel_start_time_str)  # 填开始测试时间，start time
                    start_datetime = excel_start_date.strftime("%Y-%m-%d") + " " + excel_start_time  # 先得出完整开始测试时间的字符串格式，例2023-11-1 1:2 一点2分，这是为了根据workloading推算截止日
                    start_datetime_timeformat = datetime.datetime.strptime(start_datetime,date_format)  # 开始时间字符串格式转时间格式
                    end_datetime = start_datetime_timeformat + datetime.timedelta(minutes=every_row_list[9])  # 获取完整年月日几点几分的结束日期等于加上workloading后的,结束时间是时间格式的，为了填excel，还要转字符串格式
                    end_date_excel = end_datetime.date()  # end_datetime.date()返回时间格式的年月日，可以直接填到excel,无需转为字符串
                    getplan_sheet.cell(column=6, row=every_row, value=end_date_excel)  # 填结束测试日期，end date
                    end_time_excel = (str(end_datetime.time()))[:-3]  # 虽然date_format = '%Y-%m-%d %H:%M'，end_datetime.time()还是返回的时间带有秒数，例2023-11-1 1:2:23 且为时间格式
                    # 所以再转字符串并且去除秒的部分1:2:23不要后面3位
                    getplan_sheet.cell(column=8, row=every_row, value=end_time_excel)
                    # if every_row_list[10] is not  None:
                    #     DFT_list=every_row_list[10].split('\n')
                    #     length_DFT_list=len(DFT_list)
                    #     getplan_sheet.cell(column=7, row=every_row, value='Issue verify')
                    #     getplan_sheet.cell(column=8, row=every_row, value=str(0.5*length_DFT_list))

            else:
                start_time_hour_am = start_time_hour_am_list[random.randint(0, 9)]
                if start_time_hour_am != '8':
                    excel_start_time = start_time_hour_am + ":" + start_time_hour_minutes_1  # 开始时间的字符串格式
                    excel_start_time_format = datetime.datetime.strptime(excel_start_time, date_format_1)  # 再转为时间格式
                    excel_start_time_str = excel_start_time_format.strftime(date_format_1)  # 再转为符合TCM格式的字符串
                    getplan_sheet.cell(column=7, row=every_row, value=excel_start_time_str)  # 填开始测试时间，start time
                    start_datetime = excel_start_date.strftime("%Y-%m-%d") + " " + excel_start_time  # 先得出完整开始测试时间的字符串格式，例2023-11-1 1:2 一点2分，这是为了根据workloading推算截止日
                    start_datetime_timeformat = datetime.datetime.strptime(start_datetime,date_format)  # 开始时间字符串格式转时间格式
                    end_datetime = start_datetime_timeformat + datetime.timedelta(minutes=every_row_list[9])  # 获取完整年月日几点几分的结束日期等于加上workloading后的,结束时间是时间格式的，为了填excel，还要转字符串格式
                    end_date_excel = end_datetime.date()  # end_datetime.date()返回时间格式的年月日，可以直接填到excel,无需转为字符串
                    getplan_sheet.cell(column=6, row=every_row, value=end_date_excel)  # 填结束测试日期，end date
                    end_time_excel = (str(end_datetime.time()))[:-3]  # 虽然date_format = '%Y-%m-%d %H:%M'，end_datetime.time()还是返回的时间带有秒数，例2023-11-1 1:2:23 且为时间格式
                    # 所以再转字符串并且去除秒的部分1:2:23不要后面3位
                    getplan_sheet.cell(column=8, row=every_row, value=end_time_excel)
                    # if every_row_list[10] is not  None:
                    #     DFT_list=every_row_list[10].split('\n')
                    #     length_DFT_list=len(DFT_list)
                    #     getplan_sheet.cell(column=7, row=every_row, value='Issue verify')
                    #     getplan_sheet.cell(column=8, row=every_row, value=str(0.5*length_DFT_list))

                else:
                    excel_start_time = start_time_hour_am + ":" + start_time_hour_minutes_2  # 8点开始，分钟就只能是31到59，#开始时间的字符串格式
                    excel_start_time_format = datetime.datetime.strptime(excel_start_time, date_format_1)  # 再转为时间格式
                    excel_start_time_str = excel_start_time_format.strftime(date_format_1)  # 再转为符合TCM格式的字符串
                    getplan_sheet.cell(column=7, row=every_row, value=excel_start_time_str)  # 填开始测试时间，start time
                    start_datetime = excel_start_date.strftime("%Y-%m-%d") + " " + excel_start_time  # 先得出完整开始测试时间的字符串格式，例2023-11-1 1:2 一点2分，这是为了根据workloading推算截止日
                    start_datetime_timeformat = datetime.datetime.strptime(start_datetime,date_format)  # 开始时间字符串格式转时间格式
                    end_datetime = start_datetime_timeformat + datetime.timedelta(minutes=every_row_list[9])  # 获取完整年月日几点几分的结束日期等于加上workloading后的,结束时间是时间格式的，为了填excel，还要转字符串格式
                    end_date_excel = end_datetime.date()  # end_datetime.date()返回时间格式的年月日，可以直接填到excel,无需转为字符串
                    getplan_sheet.cell(column=6, row=every_row, value=end_date_excel)  # 填结束测试日期，end date
                    end_time_excel = (str(end_datetime.time()))[:-3]  # 虽然date_format = '%Y-%m-%d %H:%M'，end_datetime.time()还是返回的时间带有秒数，例2023-11-1 1:2:23 且为时间格式
                    # 所以再转字符串并且去除秒的部分1:2:23不要后面3位
                    getplan_sheet.cell(column=8, row=every_row, value=end_time_excel)
                    # if every_row_list[10] is not  None:
                    #     DFT_list=every_row_list[10].split('\n')
                    #     length_DFT_list=len(DFT_list)
                    #     getplan_sheet.cell(column=7, row=every_row, value='Issue verify')
                    #     getplan_sheet.cell(column=8, row=every_row, value=str(0.5*length_DFT_list))


    elif start_date_format.weekday() == 6:
        excel_start_date=start_date_format + datetime.timedelta(days=1)
        if excel_start_date <= end_date_format:
            getplan_sheet.cell(column=5, row=every_row, value=excel_start_date)
            start_date_format=excel_start_date+datetime.timedelta(days=1)
            if every_row_list[9] <= 480:
                if start_time_hour_am != '8':
                    excel_start_time = start_time_hour_am + ":" + start_time_hour_minutes_1  # 开始时间的字符串格式
                    excel_start_time_format = datetime.datetime.strptime(excel_start_time, date_format_1)  # 再转为时间格式
                    excel_start_time_str = excel_start_time_format.strftime(date_format_1)  # 再转为符合TCM格式的字符串
                    getplan_sheet.cell(column=7, row=every_row, value=excel_start_time_str)  # 填开始测试时间，start time
                    start_datetime = excel_start_date.strftime("%Y-%m-%d") + " " + excel_start_time  # 先得出完整开始测试时间的字符串格式，例2023-11-1 1:2 一点2分，这是为了根据workloading推算截止日
                    start_datetime_timeformat = datetime.datetime.strptime(start_datetime,date_format)  # 开始时间字符串格式转时间格式
                    end_datetime = start_datetime_timeformat + datetime.timedelta(minutes=every_row_list[9])  # 获取完整年月日几点几分的结束日期等于加上workloading后的,结束时间是时间格式的，为了填excel，还要转字符串格式
                    end_date_excel = end_datetime.date()  # end_datetime.date()返回时间格式的年月日，可以直接填到excel,无需转为字符串
                    getplan_sheet.cell(column=6, row=every_row, value=end_date_excel)  # 填结束测试日期，end date
                    end_time_excel = (str(end_datetime.time()))[:-3]  # 虽然date_format = '%Y-%m-%d %H:%M'，end_datetime.time()还是返回的时间带有秒数，例2023-11-1 1:2:23 且为时间格式
                    # 所以再转字符串并且去除秒的部分1:2:23不要后面3位
                    getplan_sheet.cell(column=8, row=every_row, value=end_time_excel)
                    # if every_row_list[10] is not  None:
                    #     DFT_list=every_row_list[10].split('\n')
                    #     length_DFT_list=len(DFT_list)
                    #     getplan_sheet.cell(column=7, row=every_row, value='Issue verify')
                    #     getplan_sheet.cell(column=8, row=every_row, value=str(0.5*length_DFT_list))

                else:
                    excel_start_time = start_time_hour_am + ":" + start_time_hour_minutes_2  # 8点开始，分钟就只能是31到59，#开始时间的字符串格式
                    excel_start_time_format = datetime.datetime.strptime(excel_start_time, date_format_1)  # 再转为时间格式
                    excel_start_time_str = excel_start_time_format.strftime(date_format_1)  # 再转为符合TCM格式的字符串
                    getplan_sheet.cell(column=7, row=every_row, value=excel_start_time_str)  # 填开始测试时间，start time
                    start_datetime = excel_start_date.strftime("%Y-%m-%d") + " " + excel_start_time  # 先得出完整开始测试时间的字符串格式，例2023-11-1 1:2 一点2分，这是为了根据workloading推算截止日
                    start_datetime_timeformat = datetime.datetime.strptime(start_datetime,date_format)  # 开始时间字符串格式转时间格式
                    end_datetime = start_datetime_timeformat + datetime.timedelta(minutes=every_row_list[9])  # 获取完整年月日几点几分的结束日期等于加上workloading后的,结束时间是时间格式的，为了填excel，还要转字符串格式
                    end_date_excel = end_datetime.date()  # end_datetime.date()返回时间格式的年月日，可以直接填到excel,无需转为字符串
                    getplan_sheet.cell(column=6, row=every_row, value=end_date_excel)  # 填结束测试日期，end date
                    end_time_excel = (str(end_datetime.time()))[:-3]  # 虽然date_format = '%Y-%m-%d %H:%M'，end_datetime.time()还是返回的时间带有秒数，例2023-11-1 1:2:23 且为时间格式
                    # 所以再转字符串并且去除秒的部分1:2:23不要后面3位
                    getplan_sheet.cell(column=8, row=every_row, value=end_time_excel)
                    # if every_row_list[10] is not  None:
                    #     DFT_list=every_row_list[10].split('\n')
                    #     length_DFT_list=len(DFT_list)
                    #     getplan_sheet.cell(column=7, row=every_row, value='Issue verify')
                    #     getplan_sheet.cell(column=8, row=every_row, value=str(0.5*length_DFT_list))

            else:
                start_time_hour_am = start_time_hour_am_list[random.randint(0,9)]
                if start_time_hour_am != '8':
                    excel_start_time = start_time_hour_am + ":" + start_time_hour_minutes_1  # 开始时间的字符串格式
                    excel_start_time_format = datetime.datetime.strptime(excel_start_time, date_format_1)  # 再转为时间格式
                    excel_start_time_str = excel_start_time_format.strftime(date_format_1)  # 再转为符合TCM格式的字符串
                    getplan_sheet.cell(column=7, row=every_row, value=excel_start_time_str)  # 填开始测试时间，start time
                    start_datetime = excel_start_date.strftime("%Y-%m-%d") + " " + excel_start_time  # 先得出完整开始测试时间的字符串格式，例2023-11-1 1:2 一点2分，这是为了根据workloading推算截止日
                    start_datetime_timeformat = datetime.datetime.strptime(start_datetime,date_format)  # 开始时间字符串格式转时间格式
                    end_datetime = start_datetime_timeformat + datetime.timedelta(minutes=every_row_list[9])  # 获取完整年月日几点几分的结束日期等于加上workloading后的,结束时间是时间格式的，为了填excel，还要转字符串格式
                    end_date_excel = end_datetime.date()  # end_datetime.date()返回时间格式的年月日，可以直接填到excel,无需转为字符串
                    getplan_sheet.cell(column=6, row=every_row, value=end_date_excel)  # 填结束测试日期，end date
                    end_time_excel = (str(end_datetime.time()))[:-3]  # 虽然date_format = '%Y-%m-%d %H:%M'，end_datetime.time()还是返回的时间带有秒数，例2023-11-1 1:2:23 且为时间格式
                    # 所以再转字符串并且去除秒的部分1:2:23不要后面3位
                    getplan_sheet.cell(column=8, row=every_row, value=end_time_excel)
                    # if every_row_list[10] is not  None:
                    #     DFT_list=every_row_list[10].split('\n')
                    #     length_DFT_list=len(DFT_list)
                    #     getplan_sheet.cell(column=7, row=every_row, value='Issue verify')
                    #     getplan_sheet.cell(column=8, row=every_row, value=str(0.5*length_DFT_list))

                else:
                    excel_start_time = start_time_hour_am + ":" + start_time_hour_minutes_2  # 8点开始，分钟就只能是31到59，#开始时间的字符串格式
                    excel_start_time_format = datetime.datetime.strptime(excel_start_time, date_format_1)  # 再转为时间格式
                    excel_start_time_str = excel_start_time_format.strftime(date_format_1)  # 再转为符合TCM格式的字符串
                    getplan_sheet.cell(column=7, row=every_row, value=excel_start_time_str)  # 填开始测试时间，start time
                    start_datetime = excel_start_date.strftime("%Y-%m-%d") + " " + excel_start_time  # 先得出完整开始测试时间的字符串格式，例2023-11-1 1:2 一点2分，这是为了根据workloading推算截止日
                    start_datetime_timeformat = datetime.datetime.strptime(start_datetime,date_format)  # 开始时间字符串格式转时间格式
                    end_datetime = start_datetime_timeformat + datetime.timedelta(minutes=every_row_list[9])  # 获取完整年月日几点几分的结束日期等于加上workloading后的,结束时间是时间格式的，为了填excel，还要转字符串格式
                    end_date_excel = end_datetime.date()  # end_datetime.date()返回时间格式的年月日，可以直接填到excel,无需转为字符串
                    getplan_sheet.cell(column=6, row=every_row, value=end_date_excel)  # 填结束测试日期，end date
                    end_time_excel = (str(end_datetime.time()))[:-3]  # 虽然date_format = '%Y-%m-%d %H:%M'，end_datetime.time()还是返回的时间带有秒数，例2023-11-1 1:2:23 且为时间格式
                    # 所以再转字符串并且去除秒的部分1:2:23不要后面3位
                    getplan_sheet.cell(column=8, row=every_row, value=end_time_excel)
                    if every_row_list[10] is not  None:
                        DFT_list=every_row_list[10].split('\n')
                        length_DFT_list=len(DFT_list)
                        getplan_sheet.cell(column=7, row=every_row, value='Issue verify')
                        getplan_sheet.cell(column=8, row=every_row, value=str(0.5*length_DFT_list))

        else:
            excel_start_date = datetime.date(int(schedule_start[0]), int(schedule_start[1]), int(schedule_start[2]))
            getplan_sheet.cell(column=5, row=every_row, value=excel_start_date)
            start_date_format=excel_start_date+datetime.timedelta(days=1)
            if every_row_list[9] <= 480:
                if start_time_hour_am != '8':
                    excel_start_time = start_time_hour_am + ":" + start_time_hour_minutes_1  # 开始时间的字符串格式
                    excel_start_time_format = datetime.datetime.strptime(excel_start_time, date_format_1)  # 再转为时间格式
                    excel_start_time_str = excel_start_time_format.strftime(date_format_1)  # 再转为符合TCM格式的字符串
                    getplan_sheet.cell(column=7, row=every_row, value=excel_start_time_str)  # 填开始测试时间，start time
                    start_datetime = excel_start_date.strftime("%Y-%m-%d") + " " + excel_start_time  # 先得出完整开始测试时间的字符串格式，例2023-11-1 1:2 一点2分，这是为了根据workloading推算截止日
                    start_datetime_timeformat = datetime.datetime.strptime(start_datetime,date_format)  # 开始时间字符串格式转时间格式
                    end_datetime = start_datetime_timeformat + datetime.timedelta(minutes=every_row_list[9])  # 获取完整年月日几点几分的结束日期等于加上workloading后的,结束时间是时间格式的，为了填excel，还要转字符串格式
                    end_date_excel = end_datetime.date()  # end_datetime.date()返回时间格式的年月日，可以直接填到excel,无需转为字符串
                    getplan_sheet.cell(column=6, row=every_row, value=end_date_excel)  # 填结束测试日期，end date
                    end_time_excel = (str(end_datetime.time()))[:-3]  # 虽然date_format = '%Y-%m-%d %H:%M'，end_datetime.time()还是返回的时间带有秒数，例2023-11-1 1:2:23 且为时间格式
                    # 所以再转字符串并且去除秒的部分1:2:23不要后面3位
                    getplan_sheet.cell(column=8, row=every_row, value=end_time_excel)
                    # if every_row_list[10] is not  None:
                    #     DFT_list=every_row_list[10].split('\n')
                    #     length_DFT_list=len(DFT_list)
                    #     getplan_sheet.cell(column=7, row=every_row, value='Issue verify')
                    #     getplan_sheet.cell(column=8, row=every_row, value=str(0.5*length_DFT_list))

                else:
                    excel_start_time = start_time_hour_am + ":" + start_time_hour_minutes_2  # 8点开始，分钟就只能是31到59，#开始时间的字符串格式
                    excel_start_time_format = datetime.datetime.strptime(excel_start_time, date_format_1)  # 再转为时间格式
                    excel_start_time_str = excel_start_time_format.strftime(date_format_1)  # 再转为符合TCM格式的字符串
                    getplan_sheet.cell(column=7, row=every_row, value=excel_start_time_str)  # 填开始测试时间，start time
                    start_datetime = excel_start_date.strftime("%Y-%m-%d") + " " + excel_start_time  # 先得出完整开始测试时间的字符串格式，例2023-11-1 1:2 一点2分，这是为了根据workloading推算截止日
                    start_datetime_timeformat = datetime.datetime.strptime(start_datetime,date_format)  # 开始时间字符串格式转时间格式
                    end_datetime = start_datetime_timeformat + datetime.timedelta(minutes=every_row_list[9])  # 获取完整年月日几点几分的结束日期等于加上workloading后的,结束时间是时间格式的，为了填excel，还要转字符串格式
                    end_date_excel = end_datetime.date()  # end_datetime.date()返回时间格式的年月日，可以直接填到excel,无需转为字符串
                    getplan_sheet.cell(column=6, row=every_row, value=end_date_excel)  # 填结束测试日期，end date
                    end_time_excel = (str(end_datetime.time()))[:-3]  # 虽然date_format = '%Y-%m-%d %H:%M'，end_datetime.time()还是返回的时间带有秒数，例2023-11-1 1:2:23 且为时间格式
                    # 所以再转字符串并且去除秒的部分1:2:23不要后面3位
                    getplan_sheet.cell(column=8, row=every_row, value=end_time_excel)
                    # if every_row_list[10] is not  None:
                    #     DFT_list=every_row_list[10].split('\n')
                    #     length_DFT_list=len(DFT_list)
                    #     getplan_sheet.cell(column=7, row=every_row, value='Issue verify')
                    #     getplan_sheet.cell(column=8, row=every_row, value=str(0.5*length_DFT_list))

            else:
                start_time_hour_am = start_time_hour_am_list[random.randint(0,9)]
                if start_time_hour_am != '8':
                    excel_start_time = start_time_hour_am + ":" + start_time_hour_minutes_1  # 开始时间的字符串格式
                    excel_start_time_format = datetime.datetime.strptime(excel_start_time, date_format_1)  # 再转为时间格式
                    excel_start_time_str = excel_start_time_format.strftime(date_format_1)  # 再转为符合TCM格式的字符串
                    getplan_sheet.cell(column=7, row=every_row, value=excel_start_time_str)  # 填开始测试时间，start time
                    start_datetime = excel_start_date.strftime("%Y-%m-%d") + " " + excel_start_time  # 先得出完整开始测试时间的字符串格式，例2023-11-1 1:2 一点2分，这是为了根据workloading推算截止日
                    start_datetime_timeformat = datetime.datetime.strptime(start_datetime,date_format)  # 开始时间字符串格式转时间格式
                    end_datetime = start_datetime_timeformat + datetime.timedelta(minutes=every_row_list[9])  # 获取完整年月日几点几分的结束日期等于加上workloading后的,结束时间是时间格式的，为了填excel，还要转字符串格式
                    end_date_excel = end_datetime.date()  # end_datetime.date()返回时间格式的年月日，可以直接填到excel,无需转为字符串
                    getplan_sheet.cell(column=6, row=every_row, value=end_date_excel)  # 填结束测试日期，end date
                    end_time_excel = (str(end_datetime.time()))[:-3]  # 虽然date_format = '%Y-%m-%d %H:%M'，end_datetime.time()还是返回的时间带有秒数，例2023-11-1 1:2:23 且为时间格式
                    # 所以再转字符串并且去除秒的部分1:2:23不要后面3位
                    getplan_sheet.cell(column=8, row=every_row, value=end_time_excel)
                    # if every_row_list[10] is not  None:
                    #     DFT_list=every_row_list[10].split('\n')
                    #     length_DFT_list=len(DFT_list)
                    #     getplan_sheet.cell(column=7, row=every_row, value='Issue verify')
                    #     getplan_sheet.cell(column=8, row=every_row, value=str(0.5*length_DFT_list))

                else:
                    excel_start_time = start_time_hour_am + ":" + start_time_hour_minutes_2  # 8点开始，分钟就只能是31到59，#开始时间的字符串格式
                    excel_start_time_format = datetime.datetime.strptime(excel_start_time, date_format_1)  # 再转为时间格式
                    excel_start_time_str = excel_start_time_format.strftime(date_format_1)  # 再转为符合TCM格式的字符串
                    getplan_sheet.cell(column=7, row=every_row, value=excel_start_time_str)  # 填开始测试时间，start time
                    start_datetime = excel_start_date.strftime("%Y-%m-%d") + " " + excel_start_time  # 先得出完整开始测试时间的字符串格式，例2023-11-1 1:2 一点2分，这是为了根据workloading推算截止日
                    start_datetime_timeformat = datetime.datetime.strptime(start_datetime,date_format)  # 开始时间字符串格式转时间格式
                    end_datetime = start_datetime_timeformat + datetime.timedelta(minutes=every_row_list[9])  # 获取完整年月日几点几分的结束日期等于加上workloading后的,结束时间是时间格式的，为了填excel，还要转字符串格式
                    end_date_excel = end_datetime.date()  # end_datetime.date()返回时间格式的年月日，可以直接填到excel,无需转为字符串
                    getplan_sheet.cell(column=6, row=every_row, value=end_date_excel)  # 填结束测试日期，end date
                    end_time_excel = (str(end_datetime.time()))[:-3]  # 虽然date_format = '%Y-%m-%d %H:%M'，end_datetime.time()还是返回的时间带有秒数，例2023-11-1 1:2:23 且为时间格式
                    # 所以再转字符串并且去除秒的部分1:2:23不要后面3位
                    getplan_sheet.cell(column=8, row=every_row, value=end_time_excel)
                    # if every_row_list[10] is not  None:
                    #     DFT_list=every_row_list[10].split('\n')
                    #     length_DFT_list=len(DFT_list)
                    #     getplan_sheet.cell(column=7, row=every_row, value='Issue verify')
                    #     getplan_sheet.cell(column=8, row=every_row, value=str(0.5*length_DFT_list))


    else:
        excel_start_date=start_date_format
        if excel_start_date <= end_date_format:
            getplan_sheet.cell(column=5, row=every_row, value=excel_start_date)
            start_date_format=excel_start_date+datetime.timedelta(days=1)
            if every_row_list[9] <= 480:
                if start_time_hour_am != '8':
                    excel_start_time = start_time_hour_am + ":" + start_time_hour_minutes_1  # 开始时间的字符串格式
                    excel_start_time_format = datetime.datetime.strptime(excel_start_time, date_format_1)  # 再转为时间格式
                    excel_start_time_str = excel_start_time_format.strftime(date_format_1)  # 再转为符合TCM格式的字符串
                    getplan_sheet.cell(column=7, row=every_row, value=excel_start_time_str)  # 填开始测试时间，start time
                    start_datetime = excel_start_date.strftime("%Y-%m-%d") + " " + excel_start_time  # 先得出完整开始测试时间的字符串格式，例2023-11-1 1:2 一点2分，这是为了根据workloading推算截止日
                    start_datetime_timeformat = datetime.datetime.strptime(start_datetime,date_format)  # 开始时间字符串格式转时间格式
                    end_datetime = start_datetime_timeformat + datetime.timedelta(minutes=every_row_list[9])  # 获取完整年月日几点几分的结束日期等于加上workloading后的,结束时间是时间格式的，为了填excel，还要转字符串格式
                    end_date_excel = end_datetime.date()  # end_datetime.date()返回时间格式的年月日，可以直接填到excel,无需转为字符串
                    getplan_sheet.cell(column=6, row=every_row, value=end_date_excel)  # 填结束测试日期，end date
                    end_time_excel = (str(end_datetime.time()))[:-3]  # 虽然date_format = '%Y-%m-%d %H:%M'，end_datetime.time()还是返回的时间带有秒数，例2023-11-1 1:2:23 且为时间格式
                    # 所以再转字符串并且去除秒的部分1:2:23不要后面3位
                    getplan_sheet.cell(column=8, row=every_row,value=end_time_excel)  # 填结束测试时间，end time,到此为止，时间全部填完，开始时间和结束时间填的字符串格式，开始日期和结束日期填的时间格式
                    # 而且恰好结束时间是TCM标准的字符串格式
                    # if every_row_list[10] is not  None:
                    #     DFT_list=every_row_list[10].split('\n')
                    #     length_DFT_list=len(DFT_list)
                    #     getplan_sheet.cell(column=7, row=every_row, value='Issue verify')
                    #     getplan_sheet.cell(column=8, row=every_row, value=str(0.5*length_DFT_list))

                else:
                    excel_start_time = start_time_hour_am + ":" + start_time_hour_minutes_2  # 8点开始，分钟就只能是31到59，#开始时间的字符串格式
                    excel_start_time_format = datetime.datetime.strptime(excel_start_time, date_format_1)  # 再转为时间格式
                    excel_start_time_str = excel_start_time_format.strftime(date_format_1)  # 再转为符合TCM格式的字符串
                    getplan_sheet.cell(column=7, row=every_row, value=excel_start_time_str)  # 填开始测试时间，start time
                    start_datetime = excel_start_date.strftime("%Y-%m-%d") + " " + excel_start_time  # 先得出完整开始测试时间的字符串格式，例2023-11-1 1:2 一点2分，这是为了根据workloading推算截止日
                    start_datetime_timeformat = datetime.datetime.strptime(start_datetime,date_format)  # 开始时间字符串格式转时间格式
                    end_datetime = start_datetime_timeformat + datetime.timedelta(minutes=every_row_list[9])  # 获取完整年月日几点几分的结束日期等于加上workloading后的,结束时间是时间格式的，为了填excel，还要转字符串格式
                    end_date_excel = end_datetime.date()  # end_datetime.date()返回时间格式的年月日，可以直接填到excel,无需转为字符串
                    getplan_sheet.cell(column=6, row=every_row, value=end_date_excel)  # 填结束测试日期，end date
                    end_time_excel = (str(end_datetime.time()))[:-3]  # 虽然date_format = '%Y-%m-%d %H:%M'，end_datetime.time()还是返回的时间带有秒数，例2023-11-1 1:2:23 且为时间格式
                    # 所以再转字符串并且去除秒的部分1:2:23不要后面3位
                    getplan_sheet.cell(column=8, row=every_row, value=end_time_excel)
                    # if every_row_list[10] is not  None:
                    #     DFT_list=every_row_list[10].split('\n')
                    #     length_DFT_list=len(DFT_list)
                    #     getplan_sheet.cell(column=7, row=every_row, value='Issue verify')
                    #     getplan_sheet.cell(column=8, row=every_row, value=str(0.5*length_DFT_list))

            else:
                start_time_hour_am = start_time_hour_am_list[random.randint(0,9)]
                if start_time_hour_am != '8':
                    excel_start_time = start_time_hour_am + ":" + start_time_hour_minutes_1  # 开始时间的字符串格式
                    excel_start_time_format = datetime.datetime.strptime(excel_start_time, date_format_1)  # 再转为时间格式
                    excel_start_time_str = excel_start_time_format.strftime(date_format_1)  # 再转为符合TCM格式的字符串
                    getplan_sheet.cell(column=7, row=every_row, value=excel_start_time_str)  # 填开始测试时间，start time
                    start_datetime = excel_start_date.strftime("%Y-%m-%d") + " " + excel_start_time  # 先得出完整开始测试时间的字符串格式，例2023-11-1 1:2 一点2分，这是为了根据workloading推算截止日
                    start_datetime_timeformat = datetime.datetime.strptime(start_datetime,date_format)  # 开始时间字符串格式转时间格式
                    end_datetime = start_datetime_timeformat + datetime.timedelta(minutes=every_row_list[9])  # 获取完整年月日几点几分的结束日期等于加上workloading后的,结束时间是时间格式的，为了填excel，还要转字符串格式
                    end_date_excel = end_datetime.date()  # end_datetime.date()返回时间格式的年月日，可以直接填到excel,无需转为字符串
                    getplan_sheet.cell(column=6, row=every_row, value=end_date_excel)  # 填结束测试日期，end date
                    end_time_excel = (str(end_datetime.time()))[:-3]  # 虽然date_format = '%Y-%m-%d %H:%M'，end_datetime.time()还是返回的时间带有秒数，例2023-11-1 1:2:23 且为时间格式
                    # 所以再转字符串并且去除秒的部分1:2:23不要后面3位
                    getplan_sheet.cell(column=8, row=every_row, value=end_time_excel)
                    # if every_row_list[10] is not  None:
                    #     DFT_list=every_row_list[10].split('\n')
                    #     length_DFT_list=len(DFT_list)
                    #     getplan_sheet.cell(column=7, row=every_row, value='Issue verify')
                    #     getplan_sheet.cell(column=8, row=every_row, value=str(0.5*length_DFT_list))

                else:
                    excel_start_time = start_time_hour_am + ":" + start_time_hour_minutes_2  # 8点开始，分钟就只能是31到59，#开始时间的字符串格式
                    excel_start_time_format = datetime.datetime.strptime(excel_start_time, date_format_1)  # 再转为时间格式
                    excel_start_time_str = excel_start_time_format.strftime(date_format_1)  # 再转为符合TCM格式的字符串
                    getplan_sheet.cell(column=7, row=every_row, value=excel_start_time_str)  # 填开始测试时间，start time
                    start_datetime = excel_start_date.strftime("%Y-%m-%d") + " " + excel_start_time  # 先得出完整开始测试时间的字符串格式，例2023-11-1 1:2 一点2分，这是为了根据workloading推算截止日
                    start_datetime_timeformat = datetime.datetime.strptime(start_datetime,date_format)  # 开始时间字符串格式转时间格式
                    end_datetime = start_datetime_timeformat + datetime.timedelta(minutes=every_row_list[9])  # 获取完整年月日几点几分的结束日期等于加上workloading后的,结束时间是时间格式的，为了填excel，还要转字符串格式
                    end_date_excel = end_datetime.date()  # end_datetime.date()返回时间格式的年月日，可以直接填到excel,无需转为字符串
                    getplan_sheet.cell(column=6, row=every_row, value=end_date_excel)  # 填结束测试日期，end date
                    end_time_excel = (str(end_datetime.time()))[:-3]  # 虽然date_format = '%Y-%m-%d %H:%M'，end_datetime.time()还是返回的时间带有秒数，例2023-11-1 1:2:23 且为时间格式
                    # 所以再转字符串并且去除秒的部分1:2:23不要后面3位
                    getplan_sheet.cell(column=8, row=every_row, value=end_time_excel)
                    # if every_row_list[10] is not  None:
                    #     DFT_list=every_row_list[10].split('\n')
                    #     length_DFT_list=len(DFT_list)
                    #     getplan_sheet.cell(column=7, row=every_row, value='Issue verify')
                    #     getplan_sheet.cell(column=8, row=every_row, value=str(0.5*length_DFT_list))

        else:
            excel_start_date = datetime.date(int(schedule_start[0]), int(schedule_start[1]), int(schedule_start[2]))
            getplan_sheet.cell(column=5, row=every_row, value=excel_start_date)
            start_date_format=excel_start_date+datetime.timedelta(days=1)
            if every_row_list[9] <= 480:
                if start_time_hour_am != '8':
                    excel_start_time = start_time_hour_am + ":" + start_time_hour_minutes_1  # 开始时间的字符串格式
                    excel_start_time_format = datetime.datetime.strptime(excel_start_time, date_format_1)  # 再转为时间格式
                    excel_start_time_str = excel_start_time_format.strftime(date_format_1)  # 再转为符合TCM格式的字符串
                    getplan_sheet.cell(column=7, row=every_row, value=excel_start_time_str)  # 填开始测试时间，start time
                    start_datetime = excel_start_date.strftime("%Y-%m-%d") + " " + excel_start_time  # 先得出完整开始测试时间的字符串格式，例2023-11-1 1:2 一点2分，这是为了根据workloading推算截止日
                    start_datetime_timeformat = datetime.datetime.strptime(start_datetime,date_format)  # 开始时间字符串格式转时间格式
                    end_datetime = start_datetime_timeformat + datetime.timedelta(minutes=every_row_list[9])  # 获取完整年月日几点几分的结束日期等于加上workloading后的,结束时间是时间格式的，为了填excel，还要转字符串格式
                    end_date_excel = end_datetime.date()  # end_datetime.date()返回时间格式的年月日，可以直接填到excel,无需转为字符串
                    getplan_sheet.cell(column=6, row=every_row, value=end_date_excel)  # 填结束测试日期，end date
                    end_time_excel = (str(end_datetime.time()))[:-3]  # 虽然date_format = '%Y-%m-%d %H:%M'，end_datetime.time()还是返回的时间带有秒数，例2023-11-1 1:2:23 且为时间格式
                    # 所以再转字符串并且去除秒的部分1:2:23不要后面3位
                    getplan_sheet.cell(column=8, row=every_row, value=end_time_excel)
                    # if every_row_list[10] is not  None:
                    #     DFT_list=every_row_list[10].split('\n')
                    #     length_DFT_list=len(DFT_list)
                    #     getplan_sheet.cell(column=7, row=every_row, value='Issue verify')
                    #     getplan_sheet.cell(column=8, row=every_row, value=str(0.5*length_DFT_list))

                else:
                    excel_start_time = start_time_hour_am + ":" + start_time_hour_minutes_2  # 8点开始，分钟就只能是31到59，#开始时间的字符串格式
                    excel_start_time_format = datetime.datetime.strptime(excel_start_time, date_format_1)  # 再转为时间格式
                    excel_start_time_str = excel_start_time_format.strftime(date_format_1)  # 再转为符合TCM格式的字符串
                    getplan_sheet.cell(column=7, row=every_row, value=excel_start_time_str)  # 填开始测试时间，start time
                    start_datetime = excel_start_date.strftime("%Y-%m-%d") + " " + excel_start_time  # 先得出完整开始测试时间的字符串格式，例2023-11-1 1:2 一点2分，这是为了根据workloading推算截止日
                    start_datetime_timeformat = datetime.datetime.strptime(start_datetime,date_format)  # 开始时间字符串格式转时间格式
                    end_datetime = start_datetime_timeformat + datetime.timedelta(minutes=every_row_list[9])  # 获取完整年月日几点几分的结束日期等于加上workloading后的,结束时间是时间格式的，为了填excel，还要转字符串格式
                    end_date_excel = end_datetime.date()  # end_datetime.date()返回时间格式的年月日，可以直接填到excel,无需转为字符串
                    getplan_sheet.cell(column=6, row=every_row, value=end_date_excel)  # 填结束测试日期，end date
                    end_time_excel = (str(end_datetime.time()))[:-3]  # 虽然date_format = '%Y-%m-%d %H:%M'，end_datetime.time()还是返回的时间带有秒数，例2023-11-1 1:2:23 且为时间格式
                    # 所以再转字符串并且去除秒的部分1:2:23不要后面3位
                    getplan_sheet.cell(column=8, row=every_row, value=end_time_excel)
                    # if every_row_list[10] is not  None:
                    #     DFT_list=every_row_list[10].split('\n')
                    #     length_DFT_list=len(DFT_list)
                    #     getplan_sheet.cell(column=7, row=every_row, value='Issue verify')
                    #     getplan_sheet.cell(column=8, row=every_row, value=str(0.5*length_DFT_list))

            else:
                start_time_hour_am = start_time_hour_am_list[random.randint(0,9)]
                if start_time_hour_am != '8':
                    excel_start_time = start_time_hour_am + ":" + start_time_hour_minutes_1  # 开始时间的字符串格式
                    excel_start_time_format = datetime.datetime.strptime(excel_start_time, date_format_1)  # 再转为时间格式
                    excel_start_time_str = excel_start_time_format.strftime(date_format_1)  # 再转为符合TCM格式的字符串
                    getplan_sheet.cell(column=7, row=every_row, value=excel_start_time_str)  # 填开始测试时间，start time
                    start_datetime = excel_start_date.strftime("%Y-%m-%d") + " " + excel_start_time  # 先得出完整开始测试时间的字符串格式，例2023-11-1 1:2 一点2分，这是为了根据workloading推算截止日
                    start_datetime_timeformat = datetime.datetime.strptime(start_datetime,date_format)  # 开始时间字符串格式转时间格式
                    end_datetime = start_datetime_timeformat + datetime.timedelta(minutes=every_row_list[9])  # 获取完整年月日几点几分的结束日期等于加上workloading后的,结束时间是时间格式的，为了填excel，还要转字符串格式
                    end_date_excel = end_datetime.date()  # end_datetime.date()返回时间格式的年月日，可以直接填到excel,无需转为字符串
                    getplan_sheet.cell(column=6, row=every_row, value=end_date_excel)  # 填结束测试日期，end date
                    end_time_excel = (str(end_datetime.time()))[:-3]  # 虽然date_format = '%Y-%m-%d %H:%M'，end_datetime.time()还是返回的时间带有秒数，例2023-11-1 1:2:23 且为时间格式
                    # 所以再转字符串并且去除秒的部分1:2:23不要后面3位
                    getplan_sheet.cell(column=8, row=every_row, value=end_time_excel)
                    # if every_row_list[10] is not  None:
                    #     DFT_list=every_row_list[10].split('\n')
                    #     length_DFT_list=len(DFT_list)
                    #     getplan_sheet.cell(column=7, row=every_row, value='Issue verify')
                    #     getplan_sheet.cell(column=8, row=every_row, value=str(0.5*length_DFT_list))

                else:
                    excel_start_time = start_time_hour_am + ":" + start_time_hour_minutes_2  # 8点开始，分钟就只能是31到59，#开始时间的字符串格式
                    excel_start_time_format = datetime.datetime.strptime(excel_start_time, date_format_1)  # 再转为时间格式
                    excel_start_time_str = excel_start_time_format.strftime(date_format_1)  # 再转为符合TCM格式的字符串
                    getplan_sheet.cell(column=7, row=every_row, value=excel_start_time_str)  # 填开始测试时间，start time
                    start_datetime = excel_start_date.strftime("%Y-%m-%d") + " " + excel_start_time  # 先得出完整开始测试时间的字符串格式，例2023-11-1 1:2 一点2分，这是为了根据workloading推算截止日
                    start_datetime_timeformat = datetime.datetime.strptime(start_datetime,date_format)  # 开始时间字符串格式转时间格式
                    end_datetime = start_datetime_timeformat + datetime.timedelta(minutes=every_row_list[9])  # 获取完整年月日几点几分的结束日期等于加上workloading后的,结束时间是时间格式的，为了填excel，还要转字符串格式
                    end_date_excel = end_datetime.date()  # end_datetime.date()返回时间格式的年月日，可以直接填到excel,无需转为字符串
                    getplan_sheet.cell(column=6, row=every_row, value=end_date_excel)  # 填结束测试日期，end date
                    end_time_excel = (str(end_datetime.time()))[:-3]  # 虽然date_format = '%Y-%m-%d %H:%M'，end_datetime.time()还是返回的时间带有秒数，例2023-11-1 1:2:23 且为时间格式
                    # 所以再转字符串并且去除秒的部分1:2:23不要后面3位
                    getplan_sheet.cell(column=8, row=every_row, value=end_time_excel)
                    # if every_row_list[10] is not  None:
                    #     DFT_list=every_row_list[10].split('\n')
                    #     length_DFT_list=len(DFT_list)
                    #     getplan_sheet.cell(column=7, row=every_row, value='Issue verify')
                    #     getplan_sheet.cell(column=8, row=every_row, value=str(0.5*length_DFT_list))

excel.save(r'C:\excel_to_TCMAPI.xlsx')


